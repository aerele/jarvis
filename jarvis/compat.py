"""Cross-major shims for the Frappe / ERPNext APIs this app calls.

``pyproject.toml`` declares ``frappe = ">=15.0.0,<17.0.0"``, so the customer app
has to run on both majors. Three APIs moved between 15 and 16 in ways that are
not backward compatible, and each one was written against 16 and shipped onto
15 benches, where it raised at call time:

* ``File.get_content`` grew an ``encodings`` kwarg in 16. Passing it on 15 is a
  ``TypeError``, which broke every chat attachment and every ``read_file`` call.
* ``frappe.utils.xlsxutils`` was rewritten from openpyxl to xlsxwriter in 16.
  ``XLSXStyleBuilder`` does not exist on 15, and 15's ``make_xlsx`` cannot build
  a multi-tab workbook at all.
* ERPNext's ``get_itemised_tax`` takes the taxes child table on 15 and the
  parent document on 16.

Following the doctrine in :mod:`jarvis.learning.compat`, every shim probes a
capability (a signature, an importable symbol) instead of branching on
``frappe.__version__``. A version string says nothing about a backport, and the
bug this module fixes was caused by assuming one major's shape held everywhere.

Keep module-level imports to the standard library plus ``frappe``: these run in
the chat turn and tool-dispatch hot paths. Heavy optional deps (openpyxl,
xlsxwriter, erpnext) stay lazy inside the functions.
"""

import functools
import inspect

import frappe


@functools.cache
def _get_value_supports_use_local_cache() -> bool:
	"""Does ``RedisWrapper.get_value`` accept ``use_local_cache=`` (Frappe 16)?"""
	try:
		sig = inspect.signature(frappe.cache().get_value)
		return "use_local_cache" in sig.parameters
	except (AttributeError, TypeError, ValueError):
		return False


def cache_get_fresh(key):
	"""``get_value(key)`` that always reads Redis, bypassing the in-process
	``frappe.local.cache``, on Frappe 15 and 16.

	Frappe 16 has ``use_local_cache=False`` for exactly this. Frappe 15 lacks the
	kwarg (passing it is a ``TypeError``) and has no read-side local-cache bypass,
	but its ``get_value`` only consults ``local.cache`` for a key that was written
	there -- and a TTL key never is (``set_value`` with ``expires_in_sec`` skips
	local.cache on 15), provided reads pass ``expires=True`` so a miss does not
	poison it. So ``expires=True`` yields the same always-fresh read on 15.

	Callers use this only for cross-process leases/counters that must never serve a
	stale in-process copy.
	"""
	cache = frappe.cache()
	if _get_value_supports_use_local_cache():
		return cache.get_value(key, use_local_cache=False)
	return cache.get_value(key, expires=True)


def in_test() -> bool:
	"""True when running under the test runner, on Frappe 15 and 16.

	Frappe 16 exposes a module-level ``frappe.in_test``; Frappe 15 has no such
	attribute and instead sets ``frappe.flags.in_test``. Reading ``frappe.in_test``
	directly raised ``AttributeError`` on 15.

	Branch on which flag the framework actually maintains, rather than ORing both:
	on 16 the module attribute is canonical, so a stray ``flags.in_test`` must not
	widen this to True and let a scheduler-paused guard run work inline in
	production. On 15 the module attribute is absent, so the flag is authoritative.

	Call this via the module (``compat.in_test()``) so tests can patch the one
	seam on both majors; ``mock.patch("frappe.in_test", ...)`` cannot work on 15,
	where the attribute does not exist to patch.
	"""
	if hasattr(frappe, "in_test"):
		return bool(frappe.in_test)
	return bool(frappe.flags.get("in_test"))


def set_delimiters_flag(data_import_doc) -> None:
	"""Apply a Data Import's custom CSV delimiter, where the framework supports it.

	Frappe 16 added ``DataImport.set_delimiters_flag()`` (and the custom-delimiter
	feature it drives); Frappe 15 has neither and always parses with a comma.
	Calling it unconditionally was an ``AttributeError`` on 15, so skip it there:
	with no delimiter feature, the default comma parse is already correct.
	"""
	fn = getattr(data_import_doc, "set_delimiters_flag", None)
	if fn is not None:
		fn()


@functools.cache
def _get_content_takes_encodings(cls) -> bool:
	"""Does this File class accept ``get_content(encodings=...)``? (Frappe 16)"""
	try:
		return "encodings" in inspect.signature(cls.get_content).parameters
	except (AttributeError, TypeError, ValueError):
		return False


def file_bytes(fdoc) -> bytes:
	"""Raw bytes of a File document, on either Frappe major.

	Frappe 16's ``get_content(encodings=[])`` skips its text-encoding guess loop
	and hands back raw bytes. Frappe 15 has no such kwarg; its ``get_content``
	attempts a single strict ``.decode()`` and falls back to bytes on
	``UnicodeDecodeError``. That is lossless either way: a strict decode only
	succeeds on valid UTF-8, so re-encoding restores the original bytes, and a
	binary (PDF, PNG, xlsx) fails the decode and arrives as bytes untouched.
	"""
	if _get_content_takes_encodings(type(fdoc)):
		content = fdoc.get_content(encodings=[])
	else:
		content = fdoc.get_content()
	if isinstance(content, str):
		content = content.encode("utf-8", "replace")
	return content


def xlsx_bytes(sheet_data: list[tuple[str, list]]) -> bytes:
	"""Build a one-or-many-tab .xlsx workbook and return its bytes.

	Probes for ``XLSXStyleBuilder`` (the Frappe 16 rewrite) rather than for the
	``xlsxwriter`` package: xlsxwriter is a Frappe 16 dependency, but nothing
	stops it being present in a Frappe 15 virtualenv, and taking the 16 path
	there would fail in ``make_xlsx`` instead of here.
	"""
	try:
		from frappe.utils.xlsxutils import XLSXStyleBuilder
	except ImportError:
		return _xlsx_bytes_openpyxl(sheet_data)
	return _xlsx_bytes_xlsxwriter(sheet_data)


def _xlsx_bytes_xlsxwriter(sheet_data: list[tuple[str, list]]) -> bytes:
	"""Frappe 16: mirror ``make_xlsx``'s own workbook options so dates format
	identically, then let it append a bold-header worksheet per tab."""
	from io import BytesIO

	import xlsxwriter
	from frappe.utils.xlsxutils import XLSXStyleBuilder, make_xlsx

	out = BytesIO()
	wb = xlsxwriter.Workbook(
		out,
		{
			"constant_memory": True,
			"default_date_format": XLSXStyleBuilder.get_datetime_format(),
		},
	)
	for name, data in sheet_data:
		make_xlsx(data, name, wb=wb)  # adds a worksheet to `wb`, returns None
	wb.close()
	return out.getvalue()


def _xlsx_bytes_openpyxl(sheet_data: list[tuple[str, list]]) -> bytes:
	"""Frappe 15: build the workbook here instead of via ``make_xlsx``.

	15's ``make_xlsx`` saves the whole workbook on every call and returns the
	buffer, so it can only ever produce one tab: a second call against the same
	write-only workbook dies inside openpyxl's temp-file handling. It also does
	``create_sheet(name, 0)``, which prepends, so looping would reverse the
	caller's tab order. This mirrors its cell handling (bold header row, date
	and datetime number formats, html unwrapping, illegal-character stripping)
	so a workbook exported on 15 matches one exported on 16.
	"""
	import datetime
	from io import BytesIO

	import openpyxl
	from frappe.utils.xlsxutils import ILLEGAL_CHARACTERS_RE, get_excel_date_format, handle_html
	from openpyxl.cell import WriteOnlyCell
	from openpyxl.styles import Font
	from openpyxl.workbook.child import INVALID_TITLE_REGEX

	wb = openpyxl.Workbook(write_only=True)
	date_format, time_format = get_excel_date_format()

	for sheet_name, data in sheet_data:
		ws = wb.create_sheet(INVALID_TITLE_REGEX.sub(" ", sheet_name))
		ws.row_dimensions[1].font = Font(name="Calibri", bold=True)
		for row in data:
			clean_row = []
			for item in row:
				value = handle_html(item) if isinstance(item, str) else item
				if isinstance(value, str) and next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
					value = ILLEGAL_CHARACTERS_RE.sub("", value)
				if isinstance(value, datetime.date):  # datetime is a date subclass
					cell = WriteOnlyCell(ws, value=value)
					cell.number_format = (
						f"{date_format} {time_format}"
						if isinstance(value, datetime.datetime)
						else date_format
					)
					clean_row.append(cell)
				else:
					clean_row.append(value)
			ws.append(clean_row)

	out = BytesIO()
	wb.save(out)
	return out.getvalue()


def itemised_tax(doc, with_tax_account: bool = False) -> dict:
	"""Per-item tax breakup for a taxable document, on either ERPNext major.

	ERPNext 15 is ``get_itemised_tax(taxes, ...)`` and iterates the taxes child
	table. ERPNext 16 is ``get_itemised_tax(doc, ...)`` and reads
	``doc.precision()`` plus ``doc._item_wise_tax_details``.

	Dispatch on the first parameter's name, not on ``try/except TypeError``: the
	16 body raises ``TypeError`` of its own when ``_item_wise_tax_details`` is
	unset, so a blind fallback would swallow that and then fail differently on
	``doc.taxes``. The two majors also return different value shapes; both are
	passed through untouched, because the consumer is the model and inventing a
	common shape would mean fabricating fields one version does not compute.

	On 16, ``_item_wise_tax_details`` is a controller-runtime attribute, not a
	persisted field: it is set inside ``calculate_taxes_and_totals()``, which
	runs on save/submit but never on a plain ``frappe.get_doc()`` fetch. So a
	freshly loaded submitted document reaches this shim with the attribute
	unset (``None``), and passing it straight to ``get_itemised_tax`` raises
	``TypeError: 'NoneType' object is not iterable`` on every call, tax lines
	or not. Recompute it here rather than push the caller to remember to; this
	only touches the in-memory doc (no save), and is a no-op read on 15, whose
	``get_itemised_tax(doc.taxes)`` never looks at this attribute at all.

	``calculate_taxes_and_totals()`` itself early-returns without touching
	``_item_wise_tax_details`` when the document has no items at all
	(``taxes_and_totals.calculate.calculate``: ``if not len(self.doc.items):
	return``), so the attribute can still be unset afterwards - a genuinely
	taxless document, not a stale one. ``get_itemised_tax`` has no ``None``
	guard of its own, so seed the empty case directly rather than let that
	iterate a ``None`` too.

	The recompute can ALSO raise on its own, for reasons that have nothing to
	do with tax breakup: ``calculate_taxes_and_totals`` reaches
	``get_round_off_applicable_accounts(self.doc.company, ...)``, and that
	callee's ``company: str`` type annotation rejects ``None`` outright on a
	document that never had a company set - real on a bare ``frappe.new_doc()``
	(the CI bench, no default company configured; a customer's local dev bench
	with one configured never hits this, which is exactly how it first shipped
	green). A document this incomplete has nothing to report either way, so
	swallow the failure the same way missing items already degrades to `[]` -
	this dispatch's contract is "never raise on a taxless/incomplete
	document," not "compute totals correctly for one."
	"""
	from erpnext.controllers.taxes_and_totals import get_itemised_tax

	try:
		first_param = next(iter(inspect.signature(get_itemised_tax).parameters))
	except (TypeError, ValueError, StopIteration):
		first_param = ""
	# Send the child table only when the callee explicitly asks for ``taxes``;
	# anything else (the 16 name, an unreadable signature, a future rename) gets
	# the document. Defaulting the other way would hand a doc-shaped callee a
	# child table, which is the failure this shim exists to prevent.
	if first_param == "taxes":
		return get_itemised_tax(doc.get("taxes"), with_tax_account=with_tax_account)
	if doc.get("_item_wise_tax_details") is None and hasattr(doc, "calculate_taxes_and_totals"):
		try:
			doc.calculate_taxes_and_totals()
		except Exception:
			pass  # incomplete document; the None-guard below reports "no taxes"
	if doc.get("_item_wise_tax_details") is None:
		doc._item_wise_tax_details = []
	return get_itemised_tax(doc, with_tax_account=with_tax_account)


def permission_conditions(engine, doctype: str, table):
	"""Row-level permission predicate for a (possibly aliased) table, either major.

	Frappe 16 has ``Engine.get_permission_conditions(doctype, table)``, which
	builds ITS OWN role/user-permission conditions against the pypika table it
	is handed - genuinely alias-safe. But it also ANDs in this doctype's
	``permission_query_conditions`` hooks (``get_permission_query_conditions()``),
	and those return raw SQL strings hardcoded to the REAL table name (e.g.
	``ToDo``'s is literally ``` `tabToDo`.allocated_to = ... ```, byte-identical
	on 15 and 16). Handed our caller's aliased table, the 16 method ANDs that raw
	fragment straight into the outer WHERE with no protection, and MariaDB
	rejects it with "Unknown column" the moment a restricted user hits a hooked
	doctype through ``jarvis.tools.query``'s alias-everything convention. See
	``_permission_conditions_v16_with_hook`` for the fix, applied only when this
	doctype actually has such a hook (a small minority - most doctypes have none,
	and the direct 16 path stays the fast, no-subquery path for them).

	Frappe 15 has no ``Engine.get_permission_conditions`` at all: its permission
	logic lives in a separate ``Permission`` class that only does doctype-level
	``has_permission`` checks and yields no row predicate. Calling the 16 method
	on 15 raised ``AttributeError``, escaped the tool layer (the caller catches
	only ``PermissionError``) and returned HTTP 500 for every ``query`` call.

	Returns a criterion to AND into the WHERE, or ``None`` when the user is
	unrestricted. Raises ``frappe.PermissionError`` on both majors when the user
	has neither role permissions nor shared documents; the caller normalises it.
	"""
	if not hasattr(engine, "get_permission_conditions"):  # Frappe 15
		user = getattr(engine, "user", None) or frappe.session.user
		return _permission_conditions_v15(doctype, table, user)
	if _has_raw_permission_query_condition_hook(doctype):
		return _permission_conditions_v16_with_hook(engine, doctype, table)
	return engine.get_permission_conditions(doctype, table)


def _has_raw_permission_query_condition_hook(doctype: str) -> bool:
	"""True iff any installed app registers a ``permission_query_conditions``
	hook for this doctype (doctype-specific or the ``"*"`` wildcard) - the same
	lookup ``Engine.get_permission_query_conditions`` makes internally, so this
	probe sees exactly the hooks that would fire."""
	hooks = frappe.get_hooks("permission_query_conditions", {})
	return bool(hooks.get(doctype) or hooks.get("*"))


def _permission_conditions_v16_with_hook(engine, doctype: str, table):
	"""Frappe 16, doctype has a raw-SQL permission_query_conditions hook.

	Ask ``Engine.get_permission_conditions`` for the condition against the
	REAL (unaliased) table instead of the caller's ``table`` - every part of the
	result, the pypika-built role/user-permission terms AND the raw-SQL hook
	fragment, then consistently reference the real table name, so the composed
	criterion is valid SQL on its own. Scope it in a subquery over that real
	table (mirrors ``_permission_conditions_v15`` exactly, just sourcing the
	condition from the 16 engine instead of ``DatabaseQuery``), and restrict the
	caller's (possibly aliased) ``table`` by name membership - alias-safe
	either way, since ``.isin()`` never depends on the left side's alias."""
	real_table = frappe.qb.DocType(doctype)
	cond = engine.get_permission_conditions(doctype, real_table)
	if cond is None:
		return None
	sub = frappe.qb.from_(real_table).select(real_table.name).where(cond)
	return table.name.isin(sub)


def _permission_conditions_v15(doctype: str, table, user: str):
	"""Frappe 15: reuse ``DatabaseQuery.build_match_conditions``, the same engine
	``frappe.get_list`` runs on, so the composition (role permissions, the
	shared-only fallback, the if_owner constraint, User Permissions, and
	``permission_query_conditions`` hooks) is the framework's and not ours.

	It returns SQL qualified with the real table name (``tabToDo``.…), and this
	tool always aliases its tables (``FROM `tabToDo` `td```), so that fragment
	cannot go straight into the outer WHERE: MariaDB rejects it with "Unknown
	column". Rather than rewrite the framework's permission SQL, which would
	silently corrupt any hook that subqueries the same table, restrict the outer
	alias to a name set computed over the UNALIASED table. The framework's SQL is
	then correct exactly as written.
	"""
	from frappe.model.db_query import DatabaseQuery
	from pypika.terms import LiteralValue

	cond_sql = DatabaseQuery(doctype, user=user).build_match_conditions(as_condition=True)
	if not cond_sql:
		return None  # unrestricted, same as Frappe 16 returning None
	# frappe.db.sql percent-formats the statement it runs, so a LIKE pattern
	# coming from a hook has to be escaped or execution dies with "unsupported
	# format character". This is exactly what frappe's own get_match_cond does.
	inner = frappe.qb.DocType(doctype)
	sub = frappe.qb.from_(inner).select(inner.name).where(LiteralValue(cond_sql.replace("%", "%%")))
	return table.name.isin(sub)
