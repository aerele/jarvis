"""Regression tests for the Frappe 15 / 16 shims in :mod:`jarvis.compat`.

``pyproject.toml`` declares ``frappe = ">=15.0.0,<17.0.0"``, but three APIs were
written against Frappe 16 and shipped onto Frappe 15 benches, where each raised
at call time and nothing caught it:

* ``File.get_content(encodings=[])`` is a ``TypeError`` on 15. ``read_file``
  re-raised it as an unhandled HTTP 500 for every file type, and the chat attach
  path swallowed it into "[Could not read attached file ...]" for every
  attachment of every type.
* ``export_excel`` built its workbook with ``xlsxwriter`` + ``XLSXStyleBuilder``,
  neither of which exists on a 15 bench, so every export raised
  ``ModuleNotFoundError``.
* ``get_itemised_tax(doc, ...)`` is the ERPNext 16 shape; ERPNext 15 wants the
  taxes child table, so passing the document raised "'SalesInvoice' object is
  not iterable".

These tests deliberately do NOT mock ``get_content`` or the workbook builder.
Mocking them is why all three shipped: the assertions have to run against the
real Frappe installed on the bench under test, whichever major that is.

Needs DB (File doctype); FrappeTestCase rolls back per-test inserts.
"""

from __future__ import annotations

import datetime
import io
from unittest.mock import patch

import frappe
import openpyxl
from frappe.tests.utils import FrappeTestCase

from jarvis import compat
from jarvis.tools.export_excel import export_excel
from jarvis.tools.read_file import read_file


def _minimal_pdf() -> bytes:
	"""A valid one-page PDF with no text layer, built with pypdf (which
	``read_file`` already requires, so this adds no new dependency). Hand-rolling
	the bytes produces a broken xref that makes pypdf warn on every parse."""
	from pypdf import PdfWriter

	writer = PdfWriter()
	writer.add_blank_page(width=200, height=200)
	buf = io.BytesIO()
	writer.write(buf)
	return buf.getvalue()


def _make_file(name: str, payload: bytes):
	"""Insert a real private File and return the reloaded doc."""
	fdoc = frappe.get_doc(
		{
			"doctype": "File",
			"file_name": name,
			"is_private": 1,
			"content": payload,
			"decode": False,
		}
	)
	fdoc.insert(ignore_permissions=True)
	return frappe.get_doc("File", fdoc.name)


class TestFileBytesAcrossMajors(FrappeTestCase):
	"""``compat.file_bytes`` must return the exact bytes on either major."""

	def test_utf8_text_round_trips_byte_exact(self):
		payload = "hello ✓ utf8".encode()
		got = compat.file_bytes(_make_file("compat-probe.txt", payload))
		self.assertIsInstance(got, bytes)
		self.assertEqual(got, payload)

	def test_binary_is_not_corrupted(self):
		"""Frappe 15's get_content tries a strict decode first. A PDF must fail
		that decode and come back as untouched bytes, not replacement chars."""
		got = compat.file_bytes(_make_file("compat-probe.pdf", _minimal_pdf()))
		self.assertEqual(got, _minimal_pdf())
		self.assertTrue(got.startswith(b"%PDF"))

	def test_read_file_reads_a_real_attachment(self):
		"""The reported bug end to end: this raised TypeError (HTTP 500) on 15."""
		fdoc = _make_file("compat-probe-readable.txt", b"line one\nline two\n")
		out = read_file(file_url=fdoc.file_url)
		self.assertEqual(out["kind"], "text")
		self.assertIn("line two", out["text"])
		self.assertEqual(out["size_bytes"], 18)

	def test_read_file_handles_a_pdf_without_raising(self):
		fdoc = _make_file("compat-probe-readable.pdf", _minimal_pdf())
		out = read_file(file_url=fdoc.file_url)
		self.assertEqual(out["kind"], "pdf")
		# No text layer, so the tool should say so rather than blow up.
		self.assertIn("note", out)


class TestXlsxBytesAcrossMajors(FrappeTestCase):
	"""``compat.xlsx_bytes`` must produce a real, openable workbook."""

	def test_single_sheet_is_a_valid_workbook(self):
		data = [["name", "qty", "when"], ["widget", 3, datetime.date(2026, 8, 8)]]
		wb = openpyxl.load_workbook(io.BytesIO(compat.xlsx_bytes([("Alpha", data)])))
		self.assertEqual(wb.sheetnames, ["Alpha"])
		self.assertEqual([c.value for c in next(wb["Alpha"].iter_rows())], ["name", "qty", "when"])

	def test_multi_tab_preserves_caller_order(self):
		"""Frappe 15's make_xlsx does create_sheet(name, 0), which prepends, and
		saves the whole workbook per call. Looping over it would reverse the tab
		order and then die on the second save."""
		sheets = [("Alpha", [["a"], [1]]), ("Beta", [["b"], [2]]), ("Gamma", [["c"], [3]])]
		wb = openpyxl.load_workbook(io.BytesIO(compat.xlsx_bytes(sheets)))
		self.assertEqual(wb.sheetnames, ["Alpha", "Beta", "Gamma"])
		self.assertEqual([c.value for r in wb["Gamma"].iter_rows() for c in r], ["c", 3])

	def test_export_excel_tool_produces_openable_bytes(self):
		"""End to end through the tool, with save_file real, not mocked."""
		out = export_excel(rows=[{"a": 1, "b": "x"}, {"a": 2, "b": "y"}], title="compat probe")
		self.assertTrue(out["file_url"])
		fdoc = frappe.get_doc("File", out["name"])
		wb = openpyxl.load_workbook(io.BytesIO(compat.file_bytes(fdoc)))
		rows = [[c.value for c in r] for r in wb[wb.sheetnames[0]].iter_rows()]
		self.assertEqual(rows, [["a", "b"], [1, "x"], [2, "y"]])


class TestPermissionConditionsAcrossMajors(FrappeTestCase):
	"""``compat.permission_conditions`` must gate rows on either major.

	Frappe 15 has no ``Engine.get_permission_conditions``, so ``query`` returned
	HTTP 500 on every call there. ``test_query.py`` could not catch it: it does
	``patch("frappe.database.query.Engine")`` in 30 places, and a MagicMock
	answers any attribute, so it manufactures the very method whose absence is
	the bug. Those mocks are fine for asserting SQL shape, but nothing exercised
	the real permission path. These tests do, with no mocking at all, and assert
	the rowset against ``frappe.get_list`` (the framework's own answer).
	"""

	USER = "jarvis-compat-perm@example.com"

	def setUp(self):
		if not frappe.db.exists("User", self.USER):
			user = frappe.get_doc(
				{
					"doctype": "User",
					"email": self.USER,
					"first_name": "compat-perm",
					"send_welcome_email": 0,
					"user_type": "System User",
				}
			)
			user.insert(ignore_permissions=True)
		self.todos = []
		for i in range(6):
			todo = frappe.get_doc(
				{
					"doctype": "ToDo",
					"description": f"compat perm probe {i}",
					"allocated_to": self.USER if i < 2 else "Administrator",
				}
			)
			todo.insert(ignore_permissions=True)
			self.todos.append(todo.name)
		self.addCleanup(frappe.set_user, "Administrator")

	def test_real_engine_path_returns_a_usable_result(self):
		"""No mock: whatever Frappe is installed must satisfy the shim."""
		from jarvis.tools.query import _make_permission_engine

		table = frappe.qb.DocType("ToDo").as_("td")
		q = frappe.qb.from_(table).select(table.name)
		engine = _make_permission_engine(q, [table], "ToDo")
		# Administrator is unrestricted, so the predicate is None on both majors.
		self.assertIsNone(compat.permission_conditions(engine, "ToDo", table))

	def test_restricted_user_sees_exactly_what_get_list_returns(self):
		from jarvis.tools.query import query

		frappe.set_user(self.USER)
		rows = {r["name"] for r in query({"from": "ToDo", "fields": ["name"], "limit": 500})["rows"]}
		expected = {r.name for r in frappe.get_list("ToDo", fields=["name"], limit_page_length=0)}
		self.assertEqual(rows, expected)
		# and it is genuinely filtering, not returning the whole table
		self.assertLess(len(rows), frappe.db.count("ToDo"))

	def test_gate_survives_an_aliased_spec(self):
		"""ToDo's own permission_query_conditions hook emits raw
		``tabToDo``-qualified SQL (byte-identical on 15 and 16), which cannot
		resolve against ``FROM `tabToDo` `td```. Both majors' branches of
		``compat.permission_conditions`` route a hooked doctype's condition
		through a subquery on the UNALIASED table for exactly this reason - see
		``_permission_conditions_v15`` and ``_permission_conditions_v16_with_hook``."""
		from jarvis.tools.query import query

		frappe.set_user(self.USER)
		plain = {r["name"] for r in query({"from": "ToDo", "fields": ["name"], "limit": 500})["rows"]}
		aliased = query({"from": "ToDo", "alias": "td", "fields": ["td.name"], "limit": 500})["rows"]
		self.assertEqual({r["name"] for r in aliased}, plain)

	def test_hook_probe_is_selective(self):
		"""The subquery wrap (this class, above) is a real SQL-shape cost - it
		must apply ONLY to a doctype that actually has a raw-SQL
		permission_query_conditions hook (ToDo is one), not to every doctype on
		Frappe 16. No mocking: reads the real installed hooks, same lookup
		``Engine.get_permission_query_conditions`` makes internally."""
		self.assertTrue(compat._has_raw_permission_query_condition_hook("ToDo"))
		self.assertFalse(compat._has_raw_permission_query_condition_hook("Currency"))

	def test_unreadable_doctype_is_denied_cleanly(self):
		"""frappe.PermissionError from the framework must arrive as our own
		error type, not as a raw 500."""
		from jarvis.exceptions import PermissionDeniedError
		from jarvis.tools.query import query

		if not frappe.db.exists("DocType", "Sales Invoice"):
			self.skipTest("erpnext not installed on this site")
		frappe.set_user(self.USER)
		with self.assertRaises(PermissionDeniedError):
			query({"from": "Sales Invoice", "fields": ["name"], "limit": 5})


class TestItemisedTaxAcrossMajors(FrappeTestCase):
	"""``compat.itemised_tax`` must pick the argument shape this ERPNext wants."""

	def test_dispatches_without_raising_on_a_taxless_doc(self):
		if "erpnext" not in frappe.get_installed_apps():
			self.skipTest("erpnext not installed on this site")
		doc = frappe.new_doc("Sales Invoice")
		self.assertEqual(compat.itemised_tax(doc, with_tax_account=True), {})

	def test_dispatches_without_raising_on_a_freshly_fetched_doc(self):
		"""ERPNext 16's ``_item_wise_tax_details`` is a controller-runtime
		attribute set inside ``calculate_taxes_and_totals()`` (save/submit), not
		a persisted field - so it is unset (``None``) on any document object
		that was NOT just built or saved in this same Python process, which is
		exactly the shape ``get_itemised_tax_breakup`` hands this shim: a
		document loaded fresh via ``frappe.get_doc(doctype, name)``.

		``frappe.new_doc("Sales Invoice")`` (the older test above) already has
		the attribute unset, but reconstructing a doc from a plain dict is a
		closer simulation of a DB round-trip - no in-memory state survives - and
		is what actually reproduced the bug this test guards: a real call
		through ``get_itemised_tax_breakup`` raised
		``TypeError: 'NoneType' object is not iterable`` on every submitted
		invoice, taxed or not, before ``compat.itemised_tax`` recomputed the
		attribute itself."""
		if "erpnext" not in frappe.get_installed_apps():
			self.skipTest("erpnext not installed on this site")
		built = frappe.new_doc("Sales Invoice")
		fetched = frappe.get_doc(built.as_dict())
		self.assertIsNone(fetched.get("_item_wise_tax_details"))
		self.assertEqual(compat.itemised_tax(fetched, with_tax_account=True), {})

	def test_sends_the_object_the_installed_signature_asks_for(self):
		"""Guards the probe itself: ERPNext 16 names the first parameter ``doc``
		and wants the parent document, ERPNext 15 names it ``taxes`` and wants
		the child table. The spy below carries the REAL signature, so dispatch
		runs exactly as it does in production and we can see what got passed."""
		if "erpnext" not in frappe.get_installed_apps():
			self.skipTest("erpnext not installed on this site")
		import inspect

		from erpnext.controllers.taxes_and_totals import get_itemised_tax

		real_signature = inspect.signature(get_itemised_tax)
		first = next(iter(real_signature.parameters))
		self.assertIn(first, {"doc", "taxes"})

		doc = frappe.new_doc("Sales Invoice")
		doc.append("taxes", {"charge_type": "Actual", "description": "probe", "tax_amount": 0})

		seen = {}

		def spy(target, with_tax_account=False):
			seen["target"] = target
			return {}

		spy.__signature__ = real_signature
		with patch("erpnext.controllers.taxes_and_totals.get_itemised_tax", spy):
			compat.itemised_tax(doc, with_tax_account=True)

		if first == "doc":
			self.assertIs(seen["target"], doc)
		else:
			self.assertEqual(list(seen["target"]), list(doc.get("taxes")))


class TestInTestAcrossMajors(FrappeTestCase):
	"""Exercise compat.in_test()'s real body, not a mock of it. Frappe 16 keeps a
	module attribute; Frappe 15 keeps only frappe.flags.in_test."""

	def test_true_under_the_runner(self):
		# We are running under the test runner right now, on whichever major.
		self.assertTrue(compat.in_test())

	def test_module_attr_is_authoritative_and_not_widened_by_flag(self):
		"""On a major that has frappe.in_test (16), a stray flags.in_test must NOT
		flip the result - that is what would let a paused-scheduler guard run work
		inline in production."""
		if not hasattr(frappe, "in_test"):
			self.skipTest("frappe.in_test absent (Frappe 15) - covered by the fallback test")
		saved_flag = frappe.flags.get("in_test")
		try:
			with patch.object(frappe, "in_test", False):
				frappe.flags.in_test = True
				self.assertFalse(compat.in_test())  # flag must not widen it
			with patch.object(frappe, "in_test", True):
				frappe.flags.in_test = False
				self.assertTrue(compat.in_test())
		finally:
			frappe.flags.in_test = saved_flag

	def test_falls_back_to_flag_when_module_attr_absent(self):
		"""Simulate Frappe 15 (no module attribute): the flag is authoritative."""
		had = hasattr(frappe, "in_test")
		saved_attr = getattr(frappe, "in_test", None)
		saved_flag = frappe.flags.get("in_test")
		try:
			if had:
				delattr(frappe, "in_test")
			frappe.flags.in_test = True
			self.assertTrue(compat.in_test())
			frappe.flags.in_test = False
			self.assertFalse(compat.in_test())
		finally:
			if had:
				frappe.in_test = saved_attr
			frappe.flags.in_test = saved_flag


class TestSetDelimitersFlagAcrossMajors(FrappeTestCase):
	"""compat.set_delimiters_flag() calls the method on Frappe 16, where DataImport
	has it, and is a no-op on Frappe 15, where it does not exist."""

	def test_calls_the_method_when_present(self):
		class Doc:
			def __init__(self):
				self.called = False

			def set_delimiters_flag(self):
				self.called = True

		doc = Doc()
		compat.set_delimiters_flag(doc)
		self.assertTrue(doc.called)

	def test_no_op_when_absent(self):
		class Doc:
			pass

		# Must not raise on a doc lacking the v16-only method (the Frappe 15 shape).
		compat.set_delimiters_flag(Doc())


class TestCacheGetFreshAcrossMajors(FrappeTestCase):
	"""compat.cache_get_fresh() reads a TTL key straight from Redis, bypassing the
	in-process local cache, on Frappe 15 (via expires=True) and 16 (via
	use_local_cache=False). Guards the v15 local-cache poisoning fix: a cold read
	must not pin None for a key that set_value later writes."""

	def test_reads_a_ttl_key_written_before(self):
		key = "jarvis-compat-fresh-probe"
		frappe.cache().delete_value(key)
		self.addCleanup(frappe.cache().delete_value, key)
		frappe.cache().set_value(key, {"v": 1}, expires_in_sec=60)
		self.assertEqual(compat.cache_get_fresh(key), {"v": 1})

	def test_no_poison_after_cold_miss_then_set(self):
		"""The exact v15 failure shape: read (miss) -> set(expires) -> read must
		return the value, not a poisoned None."""
		key = "jarvis-compat-fresh-poison"
		frappe.cache().delete_value(key)
		self.addCleanup(frappe.cache().delete_value, key)
		self.assertIsNone(compat.cache_get_fresh(key))  # cold miss must not poison
		frappe.cache().set_value(key, {"v": 2}, expires_in_sec=60)
		self.assertEqual(compat.cache_get_fresh(key), {"v": 2})

	def test_missing_key_is_none(self):
		frappe.cache().delete_value("jarvis-compat-fresh-absent")
		self.assertIsNone(compat.cache_get_fresh("jarvis-compat-fresh-absent"))


class TestCacheGetMemoizedAcrossMajors(FrappeTestCase):
	"""compat.cache_get_memoized() keeps a TTL-key read memoized in
	frappe.local.cache for the rest of the request, without the v15 miss-poison."""

	def test_hit_is_memoized_in_local_cache(self):
		key = "jarvis-compat-memo-probe"
		cache = frappe.cache()
		cache.delete_value(key)
		self.addCleanup(cache.delete_value, key)
		cache.set_value(key, {"v": 1}, expires_in_sec=60)
		self.assertEqual(compat.cache_get_memoized(key), {"v": 1})
		# a real hit is stored back so the next read is a dict lookup
		self.assertEqual(frappe.local.cache.get(cache.make_key(key)), {"v": 1})

	def test_miss_does_not_poison(self):
		key = "jarvis-compat-memo-miss"
		cache = frappe.cache()
		cache.delete_value(key)
		self.addCleanup(cache.delete_value, key)
		self.assertIsNone(compat.cache_get_memoized(key))  # cold miss
		self.assertNotIn(cache.make_key(key), frappe.local.cache)  # not poisoned
		cache.set_value(key, {"v": 2}, expires_in_sec=60)
		self.assertEqual(compat.cache_get_memoized(key), {"v": 2})
