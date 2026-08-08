"""Read an attached file's contents so the agent can act on what the customer
sent — the read side of the artifact story (``export_excel`` / ``download_pdf``
are the write side).

The chat uploads attachments to Frappe's File doctype (private) and shows the
agent a ``📎 <name>`` marker plus the ``file_url``. The agent has no way to open
that binary on its own, so it would otherwise flail (try the browser tool on a
private URL, or hunt for a matching DocType). This tool fetches the File —
gated by the calling user's File read permission — and returns its contents in
a model-readable shape:

  - spreadsheets (xlsx / csv)  → rows (per sheet)
  - pdf                        → extracted text (per page)
  - text-like (txt/md/json/html/log/yaml/xml/tsv) → decoded text
  - images (png/jpg/…)         → metadata only. Images and PDFs the customer
                                 attaches in chat are delivered to the model as
                                 native vision at attach time, so visual content
                                 is seen directly, not through this tool.

Permission contract: the calling user must have **read** permission on the
File (Frappe's File ``has_permission`` walks the attached-doc perms for private
files). A user who can't read the file gets a clean PermissionDeniedError.
"""

import csv
import io

import frappe

from jarvis import compat
from jarvis.exceptions import InvalidArgumentError, PermissionDeniedError

_TABLE_EXT = {"xlsx", "csv", "tsv"}
_PDF_EXT = {"pdf"}
_IMAGE_EXT = {"png", "jpg", "jpeg", "gif", "webp", "bmp"}
_TEXT_EXT = {"txt", "md", "markdown", "json", "html", "htm", "log", "yaml", "yml", "xml", "csv", "tsv", "svg"}

_MAX_ROWS = 1000
_MAX_CHARS = 40000


def read_file(
	file_url: str | None = None,
	filename: str | None = None,
	sheet: str | None = None,
	max_rows: int = 500,
	max_chars: int = 20000,
) -> dict:
	"""Read an attached file and return its contents.

	Identify the file by ``file_url`` (preferred — exact) or ``filename`` (the
	name shown in the ``📎`` marker; resolves to the most recent matching File
	the user can read). ``sheet`` selects one worksheet by name for xlsx;
	omitted returns the first sheet. ``max_rows`` / ``max_chars`` bound the
	payload so a huge file doesn't blow the turn's context. Images return
	metadata only — when the customer attaches an image or PDF in chat I receive
	it as native vision and see it directly, so I don't need this tool for
	visual content (use it for spreadsheets, text, and text-PDF extraction).
	"""
	fdoc = _resolve_file(file_url, filename)
	if not frappe.has_permission("File", "read", doc=fdoc.name):
		# Generic on purpose - do not echo the resolved file's real name back
		# to a caller who isn't permitted to read it (see F12 in
		# audit-findings.md: that would confirm the existence/name of a
		# private file the caller has no other way to see).
		raise PermissionDeniedError("no matching file found")

	ext = (fdoc.file_name or "").rsplit(".", 1)[-1].lower() if "." in (fdoc.file_name or "") else ""
	# Raw bytes, on either Frappe major. Calling get_content(encodings=[])
	# directly is a TypeError on Frappe 15, which used to surface here as an
	# unhandled HTTP 500 for every file type.
	content = compat.file_bytes(fdoc)

	max_rows = min(int(max_rows or 500), _MAX_ROWS)
	max_chars = min(int(max_chars or 20000), _MAX_CHARS)

	base = {"filename": fdoc.file_name, "file_url": fdoc.file_url, "size_bytes": len(content)}

	if ext in _PDF_EXT:
		return {**base, **_read_pdf(content, max_chars)}
	if ext == "xlsx":
		return {**base, **_read_xlsx(content, sheet, max_rows)}
	if ext in ("csv", "tsv"):
		return {**base, **_read_csv(content, ext, max_rows)}
	if ext in _IMAGE_EXT:
		return {**base, **_read_image(content, ext)}
	if ext in _TEXT_EXT or _looks_text(content):
		return {**base, **_read_text(content, max_chars)}
	if ext == "xls":
		raise InvalidArgumentError("legacy .xls isn't supported; re-save as .xlsx and retry")
	return {**base, "kind": "binary", "note": f"unsupported file type {ext!r}; cannot extract text"}


def _resolve_file(file_url: str | None, filename: str | None):
	if file_url:
		name = frappe.db.get_value("File", {"file_url": file_url}, "name")
		if not name:
			raise InvalidArgumentError(f"no file found for url {file_url!r}")
		return frappe.get_doc("File", name)
	if filename:
		# Gather name-matching candidates unfiltered, then gate EACH with the
		# row-level has_permission check File actually enforces, returning the
		# first (most recent) one the caller may read. This closes the F12 hole
		# (the old code picked the most-recent match with NO permission check,
		# so a private File the caller couldn't access could be handed back)
		# while still letting a user read their OWN file: we deliberately do NOT
		# pre-filter with get_list, because File's permission-query condition is
		# a coarse, environment-sensitive approximation (attached-doc based) that
		# can exclude a caller's own unattached private file - the per-candidate
		# has_permission below is the authoritative gate.
		candidates = frappe.get_all(
			"File",
			filters={"file_name": filename},
			fields=["name"],
			order_by="creation desc",
			limit=20,
		) or frappe.get_all(
			"File",
			filters={"file_name": ["like", f"%{filename}%"]},
			fields=["name"],
			order_by="creation desc",
			limit=20,
		)
		for row in candidates:
			if frappe.has_permission("File", "read", doc=row.name):
				return frappe.get_doc("File", row.name)
		# Generic on purpose - covers both "nothing matched" and "something
		# matched but you can't read it"; the latter must not confirm the
		# existence/name of a private file the caller can't otherwise see.
		raise InvalidArgumentError("no matching file found")
	raise InvalidArgumentError("pass file_url or filename to identify the attachment")


def _read_pdf(content: bytes, max_chars: int) -> dict:
	from pypdf import PdfReader

	reader = PdfReader(io.BytesIO(content))
	pages = []
	total = 0
	for page in reader.pages:
		try:
			txt = page.extract_text() or ""
		except Exception:
			txt = ""
		pages.append(txt)
		total += len(txt)
		if total >= max_chars:
			break
	text = "\n\n".join(pages)
	result = {
		"kind": "pdf",
		"page_count": len(reader.pages),
		"text": text[:max_chars],
		"truncated": len(text) > max_chars,
	}
	if not text.strip():
		result["note"] = (
			"No extractable text; looks like a scanned PDF. If the customer "
			"attached it in chat I receive its pages as images (native vision) - "
			"read those directly instead of this tool."
		)
	return result


def _read_xlsx(content: bytes, sheet: str | None, max_rows: int) -> dict:
	import openpyxl

	wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
	names = wb.sheetnames
	targets = [sheet] if sheet and sheet in names else ([names[0]] if names else [])
	out_sheets = []
	for sname in targets:
		ws = wb[sname]
		rows = []
		for r in ws.iter_rows(values_only=True):
			rows.append([_cell(c) for c in r])
			if len(rows) >= max_rows:
				break
		out_sheets.append(
			{
				"name": sname,
				"rows": rows,
				"row_count": len(rows),
				"truncated": ws.max_row is not None and ws.max_row > len(rows),
			}
		)
	wb.close()
	return {"kind": "table", "format": "xlsx", "sheet_names": names, "sheets": out_sheets}


def _read_csv(content: bytes, ext: str, max_rows: int) -> dict:
	text = content.decode("utf-8-sig", "replace")
	delim = "\t" if ext == "tsv" else ","
	rows = []
	for row in csv.reader(io.StringIO(text), delimiter=delim):
		rows.append(row)
		if len(rows) >= max_rows:
			break
	return {
		"kind": "table",
		"format": ext,
		"sheets": [{"name": "Sheet1", "rows": rows, "row_count": len(rows)}],
	}


def _read_text(content: bytes, max_chars: int) -> dict:
	text = content.decode("utf-8", "replace")
	return {"kind": "text", "text": text[:max_chars], "truncated": len(text) > max_chars}


def _read_image(content: bytes, ext: str) -> dict:
	"""Image metadata only. Visual content reaches the model as native vision
	when the file is attached in chat (see jarvis/chat/vision.py); this tool
	deliberately does not OCR or base64-dump pixels."""
	out = {"kind": "image", "format": ext}
	try:
		from PIL import Image

		im = Image.open(io.BytesIO(content))
		out["width"], out["height"] = im.size
		out["mode"] = im.mode
		im.close()
	except Exception:
		pass
	out["note"] = (
		"Image metadata only. When the customer attaches an image or PDF in chat "
		"I receive it as native vision and can see it directly - I don't need this "
		"tool for visual content."
	)
	return out


def _looks_text(content: bytes) -> bool:
	sample = content[:2048]
	if not sample:
		return True
	if b"\x00" in sample:
		return False
	try:
		sample.decode("utf-8")
		return True
	except UnicodeDecodeError:
		return False


def _cell(v):
	import datetime
	import decimal

	if v is None:
		return ""
	if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
		return v.isoformat()
	if isinstance(v, decimal.Decimal):
		return float(v)
	return v
